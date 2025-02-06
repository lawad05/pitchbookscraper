from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.cell import MergedCell
import time

def search_company(driver, company_name):
    print(f"Searching for company: {company_name}...")
    search_bar = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "input[placeholder='Search PitchBook...']"))
    )
    search_bar.clear()
    search_bar.send_keys(company_name)
    search_bar.send_keys(Keys.RETURN)
    print("Search submitted.")
    
    try:
        first_result = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, "(//div[contains(@class, 'gs-card-item')])[1]"))
        )
        print(f"Clicking the first search result for {company_name}.")
        first_result.click()

        time.sleep(10)
        return True
    except Exception as e:
        print(f"Error finding or clicking the first search result for {company_name}: {e}")
        return False

def wrap_text(text, max_words_per_line=15):
    words = text.split()
    lines = []
    
    for i in range(0, len(words), max_words_per_line):
        lines.append(' '.join(words[i:i + max_words_per_line]))
    
    return '\n'.join(lines)


def extract_company_info(driver, company_name):
    soup = BeautifulSoup(driver.page_source, 'lxml')
    
    try:
        website = soup.find('a', href=True, target="_blank").get('href')
    except:
        website = "N/A"
    
    try:
        primary_office_heading = soup.find('span', string='Primary Office')
        if primary_office_heading:
            primary_office_li_elements = primary_office_heading.find_next('ul').find_all('li', limit=3)
            headquarters = ', '.join([item.text.strip() for item in primary_office_li_elements])
            headquarters = wrap_text(headquarters)
        else:
            headquarters = "N/A"
    except Exception as e:
        print(f"Error extracting headquarters: {e}")
        headquarters = "N/A"
    
    try:
        entity_type_heading = soup.find('span', string='Entity Types')
        if entity_type_heading:
            entity_type = entity_type_heading.find_next('div', class_='more-less-array__item more-less-array__item_77832651 more-less-array__item_d-block more-less-array__item_d-block_77832651').text.strip()
        else:
            entity_type = "N/A"
    except Exception as e:
        print(f"Error extracting entity type: {e}")
        entity_type = "N/A"

    try:
        industry_label = soup.find('span', string='Preferred Industries')
        industries = ""
        if industry_label:
            industries_container = industry_label.find_parent('div', class_='element-group__item_d2bb17bb')
            industries = industries_container.text.strip()
            industries = wrap_text(industries)
        else:
            primary_industry_label = soup.find('span', string='Primary Industry')
            if primary_industry_label:
                industry_buttons = primary_industry_label.find_next_siblings('button', class_='industry-vertical__base-indent industry-vertical__link ellipsis button button_1ac5c03e button_flat button_flat_1ac5c03e button_link button_link_1ac5c03e')
                industries = ', '.join([button.find('span', class_='button__caption button__caption_1ac5c03e').text.strip() for button in industry_buttons])
                industries = wrap_text(industries)
            else:
                industries = "N/A"
    except Exception as e:
        print(f"Error extracting industries: {e}")
        industries = "N/A"
    
    try:
        primary_contact_heading = soup.find('span', string='Primary Contact')
        if primary_contact_heading:
            contact_name = primary_contact_heading.find_next('a', href=True, target="_self").text.strip()
            contact_phone_element = primary_contact_heading.find_next('strong', string='Business:').find_next('li')
            contact_phone = contact_phone_element.text.strip().replace('Business:', '').strip() if contact_phone_element else "N/A"
            primary_contact = f"{contact_name}\n{contact_phone}"
        else:
            primary_contact = "N/A"
    except Exception as e:
        print(f"Error extracting primary contact: {e}")
        primary_contact = "N/A"

    try:
        company_overview = soup.find('p').text
        company_overview = wrap_text(company_overview)
    except:
        company_overview = "N/A"
    
    try:
        if entity_type == "Public Company":
            revenue_label = soup.find('span', string='TTM Total Revenue')
            if revenue_label:
                total_revenue = revenue_label.find_next('div', class_='element-group__item element-group__item_d2bb17bb').text.strip()
                financials = f"Total Revenue: {total_revenue}"
            else:
                financials = "N/A"
        else:
            financials = "N/A"
    except Exception as e:
        print(f"Error extracting financials: {e}")
        financials = "N/A"
    
    try:
        products_services = ', '.join([item.text for item in soup.find_all('div', class_='more-less-array__item_77832651')])
        products_services = wrap_text(products_services)
    except:
        products_services = "N/A"
    
    try:
        print("Attempting to find 'Current Team' header...")
        team_section_header = soup.find('span', string=lambda text: text and "Current Team" in text)
        if team_section_header:
            print(f"Found Current Team header: {team_section_header.text}")
            team_container = team_section_header.find_parent('div').find_next('tbody')
            if team_container:
                team_elements = team_container.find_all('a', href=True, target="_self")
                team = ', '.join([item.text.strip() for item in team_elements[:10]])
                team = wrap_text(team)
                print(f"Extracted team members: {team}")
            else:
                print("Current Team container not found.")
                team = "N/A"
        else:
            print("Current Team header not found.")
            team = "N/A"
    except Exception as e:
        print(f"Error extracting team: {e}")
        team = "N/A"
    
    try:
        employee_number_label = soup.find('span', string='Employees')
        if employee_number_label:
            employee_number = employee_number_label.find_next('p').text
        else:
            employee_number = "N/A"
    except:
        employee_number = "N/A"
    
    try:
        valuation_investors_ma = soup.find('div', {'data-test': 'profile-valuation-investors-ma'}).text
        valuation_investors_ma = wrap_text(valuation_investors_ma)
    except:
        valuation_investors_ma = "N/A"
    
    try:
        market_overview = soup.find('div', {'data-test': 'profile-market-overview'}).text
        market_overview = wrap_text(market_overview)
    except:
        market_overview = "N/A"
    
    try:
        recent_news = soup.find('div', {'data-test': 'profile-news'}).text
        recent_news = wrap_text(recent_news)
    except:
        recent_news = "N/A"

    return {
        'Company Name': company_name,
        'Entity Type': entity_type,
        'Website': website,
        'Headquarters': headquarters,
        'Industries': industries,
        'Primary Contact': primary_contact,
        'Company Overview': company_overview,
        'Financials': financials,
        'Products/Services': products_services,
        'Employee Number': employee_number,
        'Team': team,
        'Valuation/Investors/M&A Activity': valuation_investors_ma,
        'Market Overview': market_overview,
        'Recent News': recent_news
    }

def save_to_excel(all_company_info, file_name, logo_path="/Users/lucasawad/Desktop/Awad Capital Internship/Pitchbook Scraper Prototype/Logo with Tagline.jpg"):
    df = pd.DataFrame(all_company_info)
    
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col])
        except ValueError:
            continue
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{file_name} Cheatsheet"

    if logo_path:
        logo = Image(logo_path)
        logo.width = logo.width // 40
        logo.height = logo.height // 40 
        ws.add_image(logo, "A1")
    
    title = "Awad Capital Pitchbook Scraper"
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=len(df.columns))
    ws.cell(row=1, column=3).value = title
    ws.cell(row=1, column=3).font = Font(size=20, bold=True)
    ws.cell(row=1, column=3).alignment = Alignment(horizontal='left')
    
    start_row = 4
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + start_row, column=c_idx, value=value)
    
    header_fill = PatternFill(start_color="183A66", end_color="183A66", fill_type="solid") 
    text_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")   
    alternate_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
    thin_border = Border(left=Side(style='thin', color="000000"),
                         right=Side(style='thin', color="000000"),
                         top=Side(style='thin', color="000000"),
                         bottom=Side(style='thin', color="000000"))
    
    company_overview_column_letter = None
    products_services_column_letter = None 
    for r_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column)):
        for cell in row:
            if r_idx == 0:
                if cell.value == "Company Overview":
                    company_overview_column_letter = cell.column_letter
                if cell.value == "Products/Services":
                    products_services_column_letter = cell.column_letter
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center')
            else:
                if (r_idx + 1) % 2 == 0:
                    cell.fill = text_fill
                else:
                    cell.fill = alternate_fill
            cell.border = thin_border
    
    if company_overview_column_letter:
        ws.column_dimensions[company_overview_column_letter].width = 20
    if products_services_column_letter:
        ws.column_dimensions[products_services_column_letter].width = 20  
    
    for col in ws.columns:
        max_length = 0
        column_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if column_letter is None:
                column_letter = cell.column_letter 
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if column_letter and column_letter not in [company_overview_column_letter, products_services_column_letter]:
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    file_path = f"{file_name}.xlsx"
    wb.save(file_path)
    
    print(f"Data saved to {file_path}")

def main():
    company_names = input("Enter the company names (separated by commas): ").split(',')
    company_names = [name.strip() for name in company_names]
    
    file_name = input("Enter the desired name for the Excel file (without extension): ")

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_experimental_option("debuggerAddress", "localhost:9222")
    driver = webdriver.Chrome(options=chrome_options)
    print("Connecting to existing Chrome session...")
    
    print("Connected to Chrome!")
    
    all_company_info = []

    for company_name in sorted(company_names):
        try:
            success = search_company(driver, company_name)
            if not success:
                print(f"Skipping company '{company_name}' as it was not found.")
                continue

            company_info = extract_company_info(driver, company_name)

            all_company_info.append(company_info)
        except Exception as e:
            print(f"Error processing company '{company_name}': {e}")
            continue

    save_to_excel(all_company_info, file_name)

    driver.quit()

if __name__ == "__main__":
    main()

